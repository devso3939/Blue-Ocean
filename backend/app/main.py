"""Blue Ocean Opportunity Intelligence — FastAPI backend."""
from __future__ import annotations

import csv
import io
import json
import threading
import time
import uuid
from concurrent.futures import ThreadPoolExecutor
from typing import Any, Callable, Optional

from fastapi import FastAPI, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse, PlainTextResponse

from . import config, taxonomy
from .cache import cache
from .countries import get_countries
from .jobstore import JobStore
from .models import JobRequest, JobStatus
from .providers.city import CityResolutionError, CityResolver
from .services import analysis as analysis_service
from .services import opportunities as opportunities_service
from .services import snapshot as snapshot_service

app = FastAPI(title="Blue Ocean Opportunity Intelligence", version="0.1.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# Load discovered taxonomy categories from the sample (if present)
taxonomy.load_discovered()

# ---------------------------------------------------------------------------
# Job manager
# ---------------------------------------------------------------------------

class JobManager:
    def __init__(self, max_workers: int, store: Optional["JobStore"] = None):
        self._jobs: dict[str, dict[str, Any]] = {}
        self._store = store
        self._lock = threading.Lock()
        self._executor = ThreadPoolExecutor(max_workers=max_workers)
        if store is not None:
            self._recover()
            self._prune()

    # -- persistence helpers --------------------------------------------------
    def _persist(self, job_id: str, payload_json: Optional[str] = None, **fields: Any) -> None:
        if self._store is None:
            return
        try:
            if payload_json is not None:
                with self._lock:
                    job = self._jobs.get(job_id)
                    row = {
                        "job_id": job_id,
                        "kind": job["kind"] if job else "",
                        "payload": payload_json,
                        "status": fields.get("status", "queued"),
                        "stage": fields.get("stage", "queued"),
                        "progress": fields.get("progress", 0.0),
                        "message": fields.get("message"),
                        "result": fields.get("result"),
                        "error": fields.get("error"),
                        "created_at": fields.get("created_at") or (job["created_at"] if job else _now()),
                        "updated_at": fields.get("updated_at") or _now(),
                    }
                self._store.upsert(
                    row["job_id"], row["kind"], row["payload"], row["status"], row["stage"],
                    row["progress"], row["message"], row["result"], row["error"],
                    row["created_at"], row["updated_at"],
                )
            else:
                self._store.update(job_id, updated_at=_now(), **fields)
        except Exception:  # persistence must never break the job path
            pass

    def _recover(self) -> None:
        """Re-queue jobs that were queued/running when the process last stopped."""
        try:
            for row in self._store.load_all():
                result = None
                if row["status"] == "done" and row["result"]:
                    try:
                        result = json.loads(row["result"])
                    except Exception:
                        result = None
                job = {
                    "job_id": row["job_id"], "kind": row["kind"], "status": row["status"],
                    "stage": row["stage"], "progress": row["progress"], "message": row["message"],
                    "result": result, "error": row["error"],
                    "created_at": row["created_at"], "updated_at": row["updated_at"],
                }
                self._jobs[row["job_id"]] = job
                if row["status"] in ("queued", "running"):
                    payload = json.loads(row["payload"] or "{}")
                    self._jobs[row["job_id"]]["status"] = "queued"
                    self._jobs[row["job_id"]]["stage"] = "queued"
                    self._jobs[row["job_id"]]["message"] = "re-queued after restart"
                    self._executor.submit(self._run, row["job_id"], row["kind"], payload)
        except Exception:
            pass

    def _prune(self) -> None:
        try:
            import datetime
            cutoff = (datetime.datetime.now(datetime.timezone.utc)
                      - datetime.timedelta(days=7)).isoformat()
            self._store.delete_older_than(cutoff)
        except Exception:
            pass

    def submit(self, kind: str, payload: dict[str, Any]) -> str:
        job_id = uuid.uuid4().hex[:12]
        now = _now()
        with self._lock:
            self._jobs[job_id] = {
                "job_id": job_id, "kind": kind, "status": "queued",
                "stage": "queued", "progress": 0.0, "message": None,
                "result": None, "error": None, "created_at": now, "updated_at": now,
            }
        self._persist(job_id, payload_json=json.dumps(payload), status="queued",
                      stage="queued", progress=0.0, message=None, result=None,
                      error=None, created_at=now, updated_at=now)
        self._executor.submit(self._run, job_id, kind, payload)
        return job_id

    def get(self, job_id: str) -> Optional[dict[str, Any]]:
        with self._lock:
            job = self._jobs.get(job_id)
            return dict(job) if job else None

    def _run(self, job_id: str, kind: str, payload: dict[str, Any]) -> None:
        def progress(stage: str, frac: float, message: str) -> None:
            with self._lock:
                job = self._jobs.get(job_id)
                if job:
                    job["stage"] = stage
                    job["progress"] = round(min(1.0, max(0.0, frac)), 3)
                    job["message"] = message
                    job["updated_at"] = _now()
            self._persist(job_id, status="running", stage=stage,
                          progress=round(min(1.0, max(0.0, frac)), 3), message=message)

        try:
            with self._lock:
                job = self._jobs.get(job_id)
                if job:
                    job["status"] = "running"
                    job["stage"] = "started"
                    job["updated_at"] = _now()
            self._persist(job_id, status="running", stage="started")
            result = self._dispatch(kind, payload, progress)
            with self._lock:
                job = self._jobs.get(job_id)
                if job:
                    job["status"] = "done"
                    job["stage"] = "done"
                    job["progress"] = 1.0
                    job["result"] = result
                    job["updated_at"] = _now()
            self._persist(job_id, status="done", stage="done", progress=1.0,
                          message=None, error=None,
                          result=json.dumps(result, default=str))
        except Exception as e:  # noqa: BLE001
            with self._lock:
                job = self._jobs.get(job_id)
                if job:
                    job["status"] = "error"
                    job["error"] = f"{type(e).__name__}: {e}"
                    job["updated_at"] = _now()
            self._persist(job_id, status="error", error=f"{type(e).__name__}: {e}")

    def _dispatch(self, kind: str, payload: dict[str, Any], progress) -> Any:
        if kind == "resolve_city":
            country = payload.get("country", "")
            city = payload.get("city", "")
            if not city:
                raise HTTPException(400, "city is required")
            resolver = CityResolver()
            meta = resolver.resolve(country, city)
            return meta.model_dump()
        if kind == "snapshot":
            city_id = payload.get("city_id", "")
            if not city_id:
                raise HTTPException(400, "city_id is required")
            city = snapshot_service.ensure_city(city_id)
            meta = snapshot_service.get_or_build_snapshot(city, progress=progress)
            return meta.model_dump()
        if kind == "analyze":
            city_id = payload.get("city_id", "")
            category_id = payload.get("category_id", "")
            if not city_id or not category_id:
                raise HTTPException(400, "city_id and category_id are required")
            refresh = bool(payload.get("refresh", False))
            analysis = analysis_service.analyze_category(city_id, category_id, progress=progress, force=refresh)
            return analysis.model_dump()
        if kind == "opportunities":
            city_id = payload.get("city_id", "")
            if not city_id:
                raise HTTPException(400, "city_id is required")
            filters = payload.get("filters") or {}
            if payload.get("refresh"):
                filters = {**filters, "refresh": True}
            result = opportunities_service.scan_opportunities(city_id, filters, progress=progress)
            return result.model_dump()
        raise HTTPException(400, f"Unknown job kind: {kind}")


jobs = JobManager(config.JOB_MAX_WORKERS, store=JobStore(config.DATA_DIR / "jobs.sqlite"))


def _now() -> str:
    import datetime
    return datetime.datetime.now(datetime.timezone.utc).isoformat()


# ---------------------------------------------------------------------------
# Info endpoints
# ---------------------------------------------------------------------------

@app.get("/api/health")
def health():
    return {"status": "ok", "time": _now()}


@app.get("/api/config")
def api_config():
    from .providers.overture import latest_release
    return {
        "overture_release": latest_release(),
        "cache_ttls": config.TTLS,
        "peer_defaults": {
            "count": config.PEER_COUNT_DEFAULT,
            "min_count": config.PEER_MIN_COUNT,
            "range_tight": list(config.PEER_RANGE_TIGHT),
            "range_wide": list(config.PEER_RANGE_WIDE),
        },
        "scoring_weights": {"gap": config.W_GAP, "percentile": config.W_PERCENTILE, "market": config.W_MARKET},
        "disclaimer": "Estimated supply gaps are statistical market intelligence, not guaranteed demand.",
    }


@app.get("/api/countries")
def countries():
    return get_countries()


@app.get("/api/cities/search")
def cities_search(
    q: str = Query(..., min_length=2, max_length=120),
    country: str = Query("", max_length=120),
    limit: int = Query(8, ge=1, le=15),
):
    """City autocomplete via Wikidata (debounced client-side)."""
    from .countries import name_to_country_code
    from .providers import population as population_provider

    cc = name_to_country_code(country) if country else None
    try:
        candidates = population_provider.wikidata_search(q, cc, limit=limit + 4)
    except Exception as e:
        raise HTTPException(502, f"City search unavailable: {e}") from e
    import re

    city_desc = re.compile(r"city|town|capital|village|municipal|settlement|district|borough", re.I)
    out = []
    seen = set()
    for c in candidates:
        desc = c.get("description") or ""
        is_city_like = bool(c.get("is_city")) or (
            c.get("lat") is not None and c.get("lon") is not None and bool(city_desc.search(desc))
        )
        if not is_city_like:
            continue
        if cc and c.get("cc") and c["cc"].lower() != cc.lower():
            continue
        name = c.get("label") or ""
        key = (name, c.get("cc") or "")
        if key in seen or not name:
            continue
        seen.add(key)
        out.append({
            "name": name,
            "country_code": (c.get("cc") or "").upper(),
            "display_name": f"{name}, {(c.get('cc') or '').upper()}",
            "description": desc,
            "qid": c.get("qid"),
            "lat": c.get("lat"),
            "lon": c.get("lon"),
        })
    return out[:limit]


@app.get("/api/country/{cca2}/cities")
def country_cities(cca2: str):
    """Cities in a country with population, snapshot coverage and cached
    opportunity highlights — for comparing cities before drilling into one.
    """
    from .countries import country_code_to_name
    from .providers import divisions as divisions_provider
    from .providers import population as population_provider
    from .providers.city import slugify
    from .services import snapshot as snapshot_service

    cc = cca2.strip().lower()
    country_name = country_code_to_name(cca2.upper())
    # Real localities from the Overture divisions theme (same release as the
    # POI data, has populations + coordinates); Wikidata SPARQL as fallback.
    cands = divisions_provider.cities_in_country(cca2.upper(), 20_000, 50_000_000, limit=150)
    if not cands:
        qid = population_provider.country_qid(cca2.upper())
        if qid:
            try:
                cands = population_provider.peer_candidates(
                    qid, 0, 20_000, 50_000_000, limit=150
                )
            except Exception:
                cands = []
    if not cands:
        raise HTTPException(404, f"Country '{cca2}' not found")

    # Prefer English display names via the divisions' Wikidata ids (cached).
    en_names = {}
    qids = [c.get("qid") for c in cands if c.get("qid")]
    if qids:
        try:
            en_names = population_provider.english_labels(qids)
        except Exception:
            en_names = {}

    from .services.dedup import are_same_city

    out = []
    seen_qids: set[str] = set()
    seen_displays: list[str] = []  # keep for fuzzy dedup
    seen_pops: list[int] = []  # population for cross-script dedup
    for c in cands:
        cname = c.get("label") or c.get("name") or ""
        if not cname:
            continue
        en = c.get("qid") and en_names.get(c["qid"])
        display = en or cname
        pop = c.get("pop") or 0
        # Deduplicate by English name, Wikidata QID, fuzzy name match, or population proximity
        qid = c.get("qid")
        if qid and qid in seen_qids:
            continue
        display_lower = display.lower().strip()
        if any(are_same_city(display, s) for s in seen_displays):
            continue
        # Cross-script dedup: same population (within 5%) likely means same city
        if pop > 0 and any(abs(pop - sp) / max(sp, 1) < 0.05 for sp in seen_pops):
            continue
        seen_displays.append(display)
        seen_pops.append(pop)
        if qid:
            seen_qids.add(qid)
        city_id = f"{slugify(display)}-{cc}"
        row = {
            "city_id": city_id,
            "name": display,
            "population": c.get("pop"),
            "lat": c.get("lat"),
            "lon": c.get("lon"),
            "snapshot": None,
            "top_opportunities": None,
        }
        # Check snapshot with case-insensitive lookup
        meta = cache.get(f"city_snapshot:{city_id}") or cache.get(f"city_snapshot:{city_id.lower()}")
        if meta:
            total = meta.get("total_places") or 0
            pop = row["population"]
            density = round(total / pop * 10000, 1) if pop else None
            row["snapshot"] = {
                "total_places": total,
                "density_per_10k": density,
                "sparse": bool(pop and density is not None and density < config.MIN_PLACES_PER_10K),
                "fetched_at": meta.get("fetched_at"),
            }
            opps = cache.get(f"opportunities:v{config.LOGIC_VERSION}:{city_id}") or cache.get(f"opportunities:v{config.LOGIC_VERSION}:{city_id.lower()}")
            if opps:
                row["top_opportunities"] = [
                    {"label": r["label"], "score": r.get("score"), "gap": r.get("gap"), "existing": r.get("existing")}
                    for r in opps.get("opportunities", [])[:3]
                ]
        out.append(row)

    out.sort(key=lambda r: -(r.get("population") or 0))
    return {
        "country_code": cca2.upper(),
        "country_name": country_name or cca2.upper(),
        "cities": out,
    }


@app.get("/api/categories")
def categories(family: str = Query("", max_length=60), q: str = Query("", max_length=80),
               popular: bool = Query(False)):
    if q:
        return taxonomy.search_categories(q)
    if family:
        for fam in taxonomy.categories_by_family():
            if fam["id"] == family:
                return fam
        raise HTTPException(404, f"Unknown family '{family}'")
    if popular:
        return [c for c in taxonomy.all_categories() if c["popular"]]
    return taxonomy.categories_by_family()


@app.get("/api/families")
def families():
    return taxonomy.categories_by_family()


# ---------------------------------------------------------------------------
# Jobs
# ---------------------------------------------------------------------------

@app.post("/api/jobs", status_code=202)
def create_job(req: JobRequest):
    job_id = jobs.submit(req.kind, req.payload or {})
    return {"job_id": job_id, "status": "queued"}


@app.get("/api/jobs/{job_id}")
def get_job(job_id: str):
    job = jobs.get(job_id)
    if job is None:
        raise HTTPException(404, "Job not found (server may have restarted)")
    return JobStatus(**job)


# ---------------------------------------------------------------------------
# City & analysis
# ---------------------------------------------------------------------------

@app.get("/api/city/{city_id}")
def get_city(city_id: str):
    try:
        city = snapshot_service.ensure_city(city_id)
    except CityResolutionError as e:
        raise HTTPException(404, str(e)) from e
    return city.model_dump()


@app.get("/api/analysis/{analysis_id}")
def get_analysis(analysis_id: str):
    analysis = analysis_service.get_analysis(analysis_id)
    if analysis is None:
        raise HTTPException(404, "Analysis not found or expired")
    return analysis.model_dump()


@app.get("/api/opportunities/{city_id}")
def get_opportunities(city_id: str):
    data = cache.get(f"opportunities:v{config.LOGIC_VERSION}:{city_id}")
    if data is None:
        raise HTTPException(404, "No cached opportunity scan for this city. Run a scan first.")
    return data


# ---------------------------------------------------------------------------
# Export
# ---------------------------------------------------------------------------

def _place_row(p, label: str) -> list:
    gmaps = f"https://www.google.com/maps/search/?api=1&query={p.lat},{p.lon}"
    osm = f"https://www.openstreetmap.org/?mlat={p.lat}&mlon={p.lon}#map=17/{p.lat}/{p.lon}"
    return [
        p.name or "", label, p.brand or "", p.address or "",
        p.lat, p.lon, p.confidence,
        "; ".join(p.websites), "; ".join(p.phones), "; ".join(p.emails), "; ".join(p.socials),
        gmaps, osm, "; ".join(p.sources),
    ]


@app.get("/api/market")
def market_context(city_id: str = Query("", max_length=80),
                   category: str = Query("", max_length=80),
                   refresh: bool = Query(False)):
    """Country market context: World Bank indicators, buyer potential, startup estimate."""
    from .services.market import fetch_market_context
    if not city_id:
        raise HTTPException(400, "city_id is required")
    try:
        city = snapshot_service.ensure_city(city_id)
    except CityResolutionError as e:
        raise HTTPException(404, str(e)) from e
    ctx = fetch_market_context(city, category, force=refresh)
    if ctx is None:
        raise HTTPException(502, "Market data unavailable for this country")
    return ctx.model_dump()


@app.get("/api/analysis/{analysis_id}/export")
def export_analysis(analysis_id: str, format: str = Query("json", pattern="^(json|csv|xlsx)$")):
    analysis = analysis_service.get_analysis(analysis_id)
    if analysis is None:
        raise HTTPException(404, "Analysis not found or expired")
    slug = analysis.analysis_id
    if format == "json":
        path = config.EXPORT_DIR / f"analysis-{slug}.json"
        path.write_text(json.dumps(analysis.model_dump(), ensure_ascii=False, indent=2), encoding="utf-8")
        return FileResponse(str(path), media_type="application/json",
                            filename=f"analysis-{slug}.json")
    if format == "xlsx":
        return _analysis_xlsx(analysis, slug)
    # CSV
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["name", "category", "brand", "address", "lat", "lon", "confidence",
                     "website", "phone", "email", "socials", "google_maps", "openstreetmap", "source"])
    for p in analysis.places:
        writer.writerow(_place_row(p, analysis.stats.label))
    path = config.EXPORT_DIR / f"analysis-{slug}.csv"
    path.write_text(buf.getvalue(), encoding="utf-8-sig")
    return FileResponse(str(path), media_type="text/csv", filename=f"analysis-{slug}.csv")


def _analysis_xlsx(analysis, slug: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    head_fill = PatternFill("solid", fgColor="EEF2FF")
    bold = Font(bold=True)

    ws = wb.active
    ws.title = "Summary"
    s = analysis.stats
    rows = [
        ("Field", "Value"),
        ("City", f"{analysis.city.name}, {analysis.city.country}"),
        ("Population", analysis.city.population or ""),
        ("Population year", analysis.city.population_year or ""),
        ("Category", s.label),
        ("Existing businesses", s.count),
        ("Per 10,000 residents", s.per_10k),
        ("Peer benchmark (per 10k)", s.expected_per_10k),
        ("Expected at benchmark", s.expected_count),
        ("Estimated supply gap", s.gap),
        ("Gap %", s.gap_pct),
        ("Opportunity Score", s.opportunity_score),
        ("Score label", s.score_label),
        ("Data Confidence", s.data_confidence),
        ("Peers", "; ".join(f"{p.name} ({p.count})" for p in analysis.peers)),
        ("Overture release", analysis.snapshot.overture_release),
        ("POI data retrieved", analysis.snapshot.fetched_at),
        ("Disclaimer", "Estimated supply gap is statistical market intelligence, not guaranteed demand."),
    ]
    for r in rows:
        ws.append(r)
    for c in (ws["A1"], ws["B1"]):
        c.font = bold
        c.fill = head_fill
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 60

    ws2 = wb.create_sheet("Businesses")
    headers = ["name", "category", "brand", "address", "lat", "lon", "confidence",
               "website", "phone", "email", "socials", "google_maps", "openstreetmap", "source"]
    ws2.append(headers)
    for c in ws2[1]:
        c.font = bold
        c.fill = head_fill
    for p in analysis.places:
        ws2.append(_place_row(p, s.label))
    widths = [30, 16, 14, 40, 10, 10, 11, 34, 18, 26, 24, 42, 42, 20]
    for i, w in enumerate(widths, start=1):
        ws2.column_dimensions[ws2.cell(row=1, column=i).column_letter].width = w

    path = config.EXPORT_DIR / f"analysis-{slug}.xlsx"
    wb.save(path)
    return FileResponse(str(path), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        filename=f"analysis-{slug}.xlsx")


@app.get("/api/opportunities/{city_id}/export")
def export_opportunities(city_id: str, format: str = Query("csv", pattern="^(csv|xlsx)$")):
    data = cache.get(f"opportunities:v{config.LOGIC_VERSION}:{city_id}")
    if data is None:
        raise HTTPException(404, "No cached opportunity scan for this city.")
    headers = ["rank", "opportunity", "family", "existing", "per_10k", "expected",
               "gap", "gap_pct", "score", "confidence"]
    body = []
    for i, r in enumerate(data.get("opportunities", []), start=1):
        body.append([i, r["label"], r["family_label"], r["existing"], r.get("per_10k"),
                     r.get("expected"), r.get("gap"), r.get("gap_pct"),
                     r.get("score"), r.get("confidence")])
    if format == "xlsx":
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        wb = Workbook()
        ws = wb.active
        ws.title = "Opportunities"
        ws.append(headers)
        fill = PatternFill("solid", fgColor="EEF2FF")
        for c in ws[1]:
            c.font = Font(bold=True)
            c.fill = fill
        for row in body:
            ws.append(row)
        for i, w in enumerate([6, 28, 22, 10, 10, 10, 10, 10, 10, 12], start=1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
        path = config.EXPORT_DIR / f"opportunities-{city_id}.xlsx"
        wb.save(path)
        return FileResponse(str(path), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            filename=f"opportunities-{city_id}.xlsx")
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    for row in body:
        writer.writerow(row)
    path = config.EXPORT_DIR / f"opportunities-{city_id}.csv"
    path.write_text(buf.getvalue(), encoding="utf-8-sig")
    return FileResponse(str(path), media_type="text/csv",
                        filename=f"opportunities-{city_id}.csv")


# ---------------------------------------------------------------------------
# Error handlers
# ---------------------------------------------------------------------------

@app.exception_handler(CityResolutionError)
async def city_resolution_handler(request, exc: CityResolutionError):
    return JSONResponse(status_code=422, content={"detail": str(exc)})


@app.exception_handler(Exception)
async def unhandled_handler(request, exc: Exception):
    return JSONResponse(status_code=500, content={"detail": f"{type(exc).__name__}: {exc}"})


# ---------------------------------------------------------------------------
# Serve frontend static files (for Render / Docker deployment)
# ---------------------------------------------------------------------------
import os
from pathlib import Path
from fastapi.staticfiles import StaticFiles
from starlette.responses import FileResponse as StarletteFileResponse

_frontend_dir = Path(os.environ.get("BLUEOCEAN_FRONTEND_DIR", "/app/frontend_out"))
if _frontend_dir.is_dir() and (_frontend_dir / "index.html").exists():
    # Mount Next.js static assets first (most specific)
    _static = _frontend_dir / "_next" / "static"
    if _static.is_dir():
        app.mount("/_next/static", StaticFiles(directory=str(_static)), name="next-static")

    _index_html = _frontend_dir / "index.html"

    @app.get("/{full_path:path}", include_in_schema=False)
    async def serve_spa(full_path: str):
        """SPA catch-all: serve the file if it exists, otherwise index.html."""
        # Try to serve the exact file first (favicon.ico, manifest.json, etc.)
        file_path = _frontend_dir / full_path
        if file_path.is_file():
            return StarletteFileResponse(str(file_path))
        # For everything else (dynamic routes), serve index.html
        return StarletteFileResponse(str(_index_html))
